from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl , xlrd
from openpyxl import Workbook
import pathlib
from tkinter import StringVar
from tkinter import PhotoImage
from fpdf import FPDF

background = "#d5e5f0"
# framebg = "#e87c2e"
# framefg = "#0f4881"

root = Tk()
root.title("student registration system")
root.geometry('1250x700+250+100')
root.config(bg=background)


file = pathlib.Path('xlsx/student_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet.title = 'Student Data'
    sheet['A1'] = 'Registration No.'
    sheet['B1'] = 'Name'
    sheet['C1'] = 'Dob'
    sheet['D1'] = 'Gender'
    sheet['E1'] = 'Course'
    sheet['F1'] = 'Year'
    sheet['G1'] = 'Mobile No.'
    sheet['H1'] = 'Email'
    sheet['I1'] = 'Address'
    sheet['J1'] = 'Father Name'
    sheet['K1'] = 'Mother Name'
    sheet['L1'] = 'date of registration'
   

    file.save('xlsx/student_data.xlsx')


# exit window
def Exit():
    root.destroy()

#search
def search():
    global x4
    text = Search.get()
    clear()
    # saveButton.config(state='disabled')
    # print(text)
    try:
        int(text)
    except:
        messagebox.showerror("error", "Invalid Registration Number !")
        return

    file=openpyxl.load_workbook("xlsx/student_data.xlsx")
    sheet=file.active
    for row in sheet.rows:
        # print(row[0].value)
   
        if row[0].value == text:
           #print(row[0].value)
           search_row = row
           break

    x1 = search_row[0].value
    x2 = search_row[1].value
    x3 = search_row[2].value
    x4 = search_row[3].value # gender
    x5 = search_row[4].value
    x6 = search_row[5].value
    x7 = search_row[6].value
    x8 = search_row[7].value
    x9 = search_row[8].value
    x10 = search_row[9].value
    x11 = search_row[10].value
    x12 = search_row[11].value

    # print(x1, x2, x3, x4, x5, x6, x7, x8, x9, x10, x11, x12)
    Registration.set(x1)
    Name.set(x2)
    DOB.set(x3)
    if x4 == 'Male':
        radio.set(1)
    else:
        radio.set(2)
    course_entry.set(x5)
    Year.set(x6)
    Mobile.set(x7)
    Email.set(x8)
    Address.set(x9)
    Father_Name.set(x10)
    Mother_Name.set(x11)
    Date.set(x12)


    img = (Image.open("student_images/"+str(x1)+".jpg"))
    resize_image = img.resize((190, 190))
    photo =ImageTk.PhotoImage(resize_image)
    image_label.config(image=photo)
    image_label.image = photo

#update
def update():
    selection()
    R1 = Registration.get()
    N1 = Name.get()
    D1 = DOB.get()
    G1 = gender
    C1 = course_entry.get()
    Y1 = Year.get()
    M1 = Mobile.get()
    E1 = Email.get()
    A1 = Address.get()
    F1 = Father_Name.get()
    Mo1 = Mother_Name.get()

    # print(R1)
    file = openpyxl.load_workbook("xlsx/student_data.xlsx")
    sheet = file.active

    for row in sheet.rows:
        # print(row[0].value)

        if row[0].value == str(R1):
            # print(row[0].value)
            update_row = row
            # print(update_row)
            break
    
    update_row[1].value = N1
    update_row[2].value = D1
    update_row[3].value = G1
    update_row[4].value = C1
    update_row[5].value = Y1
    update_row[6].value = M1
    update_row[7].value = E1
    update_row[8].value = A1
    update_row[9].value = F1
    update_row[10].value = Mo1

    # sheet.cell(column=2, row=update_row, value=N1)
    # sheet.cell(column=3, row=update_row, value=D1)
    # sheet.cell(column=4, row=update_row, value=G1)
    # sheet.cell(column=5, row=update_row, value=C1)
    # sheet.cell(column=6, row=update_row, value=Y1)
    # sheet.cell(column=7, row=update_row, value=M1)
    # sheet.cell(column=8, row=update_row, value=E1)
    # sheet.cell(column=9, row=update_row, value=A1)
    # sheet.cell(column=10, row=update_row, value=F1)
    # sheet.cell(column=11, row=update_row, value=Mo1)
  
    file.save('xlsx/student_data.xlsx')

    try:
        img.save("student_images/"+str(R1)+".jpg")
    except:
        pass

    messagebox.showinfo("Update", "Update Successfully !")

    clear()




def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
        print(gender)
    else:
        gender = "Female"
        print(gender)

def showimage():
    global img
    global filename
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select Image File", filetypes=(("JPG File", "*.jpg"))) #, ("PNG File", "*.png"), ("All Files", "*.txt")))
    img = Image.open(filename)
    resized_image = img.resize((190, 190))
    imgtk = ImageTk.PhotoImage(resized_image)
    image_label.config(image=imgtk)
    image_label.image = imgtk


#registration no
def registraion_no():
    file = openpyxl.load_workbook('xlsx/student_data.xlsx')
    sheet = file.active
    row = sheet.max_row
    max_row_value = sheet.cell(row=row, column=1).value
    try:
        Registration.set(int(max_row_value)+1)
    except:
        Registration.set("1")

def clear():
    global img
    Name.set('')
    DOB.set('')
    Year.set('')
    Mobile.set('')
    Email.set('')
    Address.set('')
    Father_Name.set('')
    Mother_Name.set('')
    course_entry.set("Select Course")

    registraion_no()

    
    # img1 = PhotoImage(file="images/user.png")
    img1 =ImageTk.PhotoImage(Image.open("images/user.jpg"))
    image_label.config(image=img1)
    image_label.image = img1

    img = ""

    # saveButton.config(state = 'normal')
    print('CLEAR')

def save():
    global img
    R1 = Registration.get()
    N1 = Name.get()
    D1 = DOB.get()
    Y1 = Year.get()
    M1 = Mobile.get()
    E1 = Email.get()
    A1 = Address.get()
    F1 = Father_Name.get()
    Mo1 = Mother_Name.get()
    C1 = course_entry.get()
    try:
        G1 = gender
    except:
        messagebox.showerror("error", "Select Gender !")
        return
    
    if N1 == '' or D1 == '' or G1 == '' or C1 == '' or Mo1 == '' or Y1 == '' or M1 == '' or E1 == '' or A1 =='' or F1 == '':
        messagebox.showerror("error", "Something is missing !")
    else:
        file = openpyxl.load_workbook('xlsx/student_data.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=D1)
        sheet.cell(column=4, row=sheet.max_row, value=G1)
        sheet.cell(column=5, row=sheet.max_row, value=C1)
        sheet.cell(column=6, row=sheet.max_row, value=Y1)
        sheet.cell(column=7, row=sheet.max_row, value=M1)
        sheet.cell(column=8, row=sheet.max_row, value=E1)
        sheet.cell(column=9, row=sheet.max_row, value=A1)
        sheet.cell(column=10, row=sheet.max_row, value=F1)
        sheet.cell(column=11, row=sheet.max_row, value=Mo1)
        sheet.cell(column=12, row=sheet.max_row, value=date.today())
   

        file.save('xlsx/student_data.xlsx')

        try:
            img.save("student_images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info", "Image is not available !")

        messagebox.showinfo("info", "Saved Successfully !")

        clear()
        registraion_no()

        print('SAVE')

# def generate_pdf():
#     file = openpyxl.load_workbook('xlsx/student_data.xlsx')
#     sheet = file.active
#     # sheet = wb['Sheet1']

#     pdf = FPDF()
#     pdf.add_page()
#     pdf.set_font("Arial", size=15)

#     for row in range(2, sheet.max_row + 1):
#         name = sheet.cell(row=row, column=1).value
#         email = sheet.cell(row=row, column=2).value
#         phone = sheet.cell(row=row, column=3).value

#         pdf.cell(200, 10, txt=f"Name: {name}", ln=True, align='L')
#         pdf.cell(200, 10, txt=f"Email: {email}", ln=True, align='L')
#         pdf.cell(200, 10, txt=f"Phone: {phone}", ln=True, align='L')
#         pdf.ln(10)

#     pdf.output("application.pdf")
    # status_label.config(text="PDF generated successfully", fg="green")



# top frames
Label(root, text="Email : psamanta1401@gmail.com", width=10, height=3, bg="#5083eb", anchor='e').pack(side=TOP, fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=2, bg="#0f4881", fg="white", font='arial 20 bold').pack(side=TOP, fill=X)
    
# search box update
Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font='arial 20').place(x=900, y=70)
# imageicon3 = PhotoImage(file="images/search.png")
Srch = Button(root, text="Search", width=6, height=1, fg='white', bg='#0f4881', font='arial 14 bold',command=search)
Srch.place(x=1150, y=70)

# imageicon4 = PhotoImage(file="images/update.png")
update_button = Button(root,text="Update", width=6, height=1, fg='white',bg="#0f4881",font='arial 14 bold',command=update)
update_button.place(x=200, y=70)

exit_button = Button(root,text="Exit", width=6, height=1, fg='white',bg="#0f4881",font='arial 14 bold', command=Exit)
exit_button.place(x=110, y=70)


# Registration and Date
Label(root, text="Registration No:", width=20, height=3, bg=background).place(x=30, y=135)
Label(root, text="Date:", width=10, height=2, bg=background).place(x=520, y=140)
Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=35, font='arial 10')
reg_entry.place(x=160, y=150)



registraion_no()



today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font='arial 10')
date_entry.place(x=600, y=150)
Date.set(d1)

# student details
obj = Label(root, text="Student's details :", width=100, bd=2, height=17, bg=background,font=20, relief=GROOVE , anchor="n")
obj.place(x=30, y=200)

Label(obj, text="Full Name:", bg=background).place(x=30, y=50)
Label(obj, text="DOB:",  bg=background).place(x=30, y=100)
Label(obj, text="Gender:", bg=background).place(x=30, y=150)
Label(obj, text="Course:",  bg=background).place(x=30, y=200)
Label(obj, text="Year:",  bg=background).place(x=30, y=250)

Label(obj, text="Mobile No:", bg=background).place(x=500, y=50)
Label(obj, text="Email:", bg=background).place(x=500, y=100)
Label(obj, text="Address:",  bg=background).place(x=500, y=150)
Label(obj, text="Father's Name:",  bg=background).place(x=500, y=200)
Label(obj, text="Mother's Name:",  bg=background).place(x=500, y=250)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font='arial 10')
name_entry.place(x=160, y=50)

DOB = StringVar()
dob_entry = Entry(obj, textvariable=DOB, width=20, font='arial 10')
dob_entry.place(x=160, y=100)

radio = IntVar()
radio1 = Radiobutton(obj, text="Male", variable=radio, value=1, command=selection)
radio1.place(x=160, y=150)
radio2 = Radiobutton(obj, text="Female", variable=radio, value=2, command=selection)
radio2.place(x=200, y=150)


course_entry = Combobox(obj, values=['BSC','BTECH','MSC','MTECH'], width=20, font='arial 10', state="r")
course_entry.place(x=160, y=200)
course_entry.set("Select Course")

Year = IntVar()
year_entry = Entry(obj, textvariable=Year, width=20, font='arial 10')
year_entry.place(x=160, y=250)

Mobile = IntVar()
mobile_entry = Entry(obj, textvariable=Mobile, width=20, font='arial 10')
mobile_entry.place(x=630, y=50)

Email = StringVar()
email_entry = Entry(obj, textvariable=Email, width=20, font='arial 10')
email_entry.place(x=630, y=100)

Address = StringVar()
address_entry = Entry(obj, textvariable=Address, width=20, font='arial 10')
address_entry.place(x=630, y=150)

Father_Name = StringVar()
father_entry = Entry(obj, textvariable=Father_Name, width=20, font='arial 10')
father_entry.place(x=630, y=200)

Mother_Name = StringVar()
mother_entry = Entry(obj, textvariable=Mother_Name, width=20, font='arial 10')
mother_entry.place(x=630, y=250)

# image
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

image =ImageTk.PhotoImage(Image.open("images/user.jpg"))
image_label = Label(f, image=image, bg="black")
image_label.place(x=0, y=0)

#button
Button(root, text="Upload", width=12, height=1, font='arial 12 bold', bg="#0f4881", fg='white',command=showimage).place(x=1000, y=370)
saveButton = Button(root, text="Save", state='normal',width=12, height=1, font='arial 12 bold', bg="#0f4881", fg='white',command=save).place(x=1000, y=410)
resetButton=Button(root, text="Reset", width=12, height=1, font='arial 12 bold', bg="#0f4881", fg='white',command=clear).place(x=1000, y=450)
deletButton = Button(root, text="Delete", width=12, height=1, font='arial 12 bold', bg="#0f4881", fg='white').place(x=1000, y=490)




root.mainloop()

